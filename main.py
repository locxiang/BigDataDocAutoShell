"""主程序入口 - 文档自动提取关键信息系统"""
import sys
import logging
import time
import threading
from pathlib import Path
from datetime import datetime

from src.config import validate_config, DATA_DIR, OUTPUT_DIR
from src.document_reader import DocumentReader
from src.document_classifier import DocumentClassifier
from src.information_extractor import InformationExtractor
from src.data_storage import DataStorage
from src.display import Display

# 配置日志 - 只写入文件，不输出到控制台
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('processing.log', encoding='utf-8')
    ]
)
logger = logging.getLogger(__name__)


class DocumentProcessor:
    """文档处理器"""
    
    def __init__(self):
        """初始化处理器"""
        self.reader = DocumentReader()
        self.classifier = DocumentClassifier()
        self.extractor = InformationExtractor()
        self.storage = DataStorage()
        self.display = Display(log_lines=10)
        
        # 统计信息
        self.stats = {
            'total': 0,
            'success': 0,
            'failed': 0,
            'failed_files': [],
            'start_time': None,
            'end_time': None,
        }
    
    def print_header(self):
        """打印启动信息"""
        print("=" * 50)
        print("文档自动提取关键信息系统")
        print("=" * 50)
        print(f"数据目录: {DATA_DIR}")
        print(f"输出目录: {OUTPUT_DIR}")
        print(f"开始时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("=" * 50)
        print()
    
    def update_status(self, index: int, total: int, status: str, file_name: str, extra: str = ""):
        """
        更新处理状态（显示在界面上）
        
        Args:
            index: 当前文件索引（从1开始）
            total: 总文件数
            status: 状态（读取、分类、提取、保存、成功、失败）
            file_name: 文件名
            extra: 额外信息
        """
        status_map = {
            '读取': '📖',
            '分类': '🏷️',
            '提取': '🔍',
            '保存': '💾',
            '成功': '✓',
            '失败': '✗',
        }
        status_icon = status_map.get(status, f'[{status}]')
        
        message = f"{status_icon} {file_name}"
        if extra:
            message += f" {extra}"
        
        # 判断是否为错误/失败消息
        is_error = (status == '失败')
        
        # 添加到日志缓冲区（失败消息显示为黄色）
        self.display.add_log(message, is_error=is_error)
        
        # 记录到日志文件
        logger.info(f"[{index}/{total}] {status} {file_name} {extra}")
    
    def process_file(self, file_path: Path, index: int, total: int) -> bool:
        """
        处理单个文件
        
        Args:
            file_path: 文件路径
            index: 文件索引
            total: 总文件数
            
        Returns:
            是否处理成功
        """
        file_name = file_path.name
        
        try:
            # 1. 读取文档
            self.update_status(index, total, '读取', file_name)
            self._render_display(index, total, file_name, "📖 正在读取文档...")
            content = self.reader.read_document(file_path)
            if not content:
                raise ValueError("无法读取文档内容")
            
            # 预处理文本
            content = self.reader.preprocess_text(content)
            
            # 2. 文档分类
            self.update_status(index, total, '分类', file_name)
            self._render_display(index, total, file_name, "🏷️ 正在分类文档...")
            doc_type = self.classifier.classify(content, file_name)
            if not doc_type:
                raise ValueError("文档分类失败")
            
            # 3. 信息提取
            self.update_status(index, total, '提取', file_name)
            self._render_display(index, total, file_name, "🔍 正在提取信息...")
            data = self.extractor.extract(content, doc_type, file_name)
            if not data:
                raise ValueError("信息提取失败")
            
            # 数据清洗
            data = self.extractor.clean_data(data)
            
            # 4. 保存数据
            self.update_status(index, total, '保存', file_name)
            self._render_display(index, total, file_name, "💾 正在保存数据...")
            success = self.storage.save_data(data, doc_type)
            if not success:
                raise ValueError("数据保存失败")
            
            # 5. 成功
            self.update_status(index, total, '成功', file_name, f"→ {doc_type}")
            self._render_display(index, total, file_name, f"✓ 处理成功 → {doc_type}")
            return True
            
        except Exception as e:
            error_msg = str(e)
            logger.error(f"处理文件失败: {file_name}, 错误: {error_msg}")
            
            # 简化错误信息用于显示
            simplified_error = self._simplify_error(error_msg, file_name)
            self.update_status(index, total, '失败', file_name, f"- {simplified_error}")
            
            # 保存完整错误信息到统计中
            self.stats['failed_files'].append({
                'file': file_name,
                'error': error_msg,
                'simplified_error': simplified_error
            })
            self._render_display(index, total, file_name, f"✗ 处理失败 - {simplified_error}")
            return False
    
    def _simplify_error(self, error_msg: str, file_name: str) -> str:
        """
        简化错误信息，提取关键错误原因
        
        Args:
            error_msg: 完整错误信息
            file_name: 文件名
            
        Returns:
            简化后的错误信息
        """
        # 提取文件后缀
        file_ext = Path(file_name).suffix if file_name else ""
        
        # 简化错误信息
        # 移除重复的文件名和路径信息
        simplified = error_msg
        
        # 移除文件路径
        import re
        simplified = re.sub(r'/Users/[^:]+', '', simplified)
        simplified = re.sub(r'data/[^:]+', '', simplified)
        
        # 移除重复的文件名
        simplified = re.sub(re.escape(file_name) + r'[：:]\s*', '', simplified)
        
        # 提取关键错误原因
        # 常见错误模式
        if 'is not a Word Document' in simplified:
            return '文件格式错误（不是有效的Word文档）'
        elif 'antiword执行失败' in simplified or 'antiword' in simplified.lower():
            if 'is not a Word Document' in error_msg:
                return '文件格式错误（不是有效的Word文档）'
            else:
                return 'antiword读取失败'
        elif '无法读取' in simplified:
            # 提取具体原因
            if '文档内容' in simplified:
                return '无法读取文档内容'
            else:
                return '文件读取失败'
        elif '超时' in simplified:
            return '读取超时'
        elif '未安装' in simplified:
            return '依赖工具未安装'
        else:
            # 提取最后的关键信息（去除多余描述）
            parts = simplified.split('：')
            if len(parts) > 1:
                return parts[-1].strip()[:50]  # 限制长度
            else:
                return simplified.strip()[:50]  # 限制长度
    
    def _render_display(self, index: int, total: int, current_file: str, status: str = ""):
        """渲染显示界面"""
        from datetime import datetime
        
        # 更新显示统计信息
        self.display.update_stats(
            index=index,
            total=total,
            success=self.stats['success'],
            failed=self.stats['failed'],
            start_time=datetime.fromtimestamp(self.stats['start_time']) if self.stats['start_time'] else None
        )
        
        # 渲染显示
        self.display.render([], current_file, status)
    
    def print_summary(self):
        """打印统计信息（已在处理完成后通过 Textual UI 显示，此方法保留用于兼容性）"""
        # 总结信息已经在处理完成后通过 SummaryScreen 显示
        # 此方法保留用于兼容性，但不再需要执行任何操作
        pass
    
    def _process_files_in_thread(self, files):
        """在后台线程中处理文件"""
        try:
            self.stats['total'] = len(files)
            self.stats['start_time'] = time.time()
            
            # 初始化统计信息显示
            from datetime import datetime
            self.display.update_stats(
                index=0,
                total=self.stats['total'],
                success=0,
                failed=0,
                start_time=datetime.fromtimestamp(self.stats['start_time'])
            )
            
            self.display.add_log(f"开始处理 {self.stats['total']} 个文件")
            
            # 处理每个文件
            for index, file_path in enumerate(files, 1):
                success = self.process_file(file_path, index, len(files))
                
                if success:
                    self.stats['success'] += 1
                else:
                    self.stats['failed'] += 1
            
            # 记录结束时间
            self.stats['end_time'] = time.time()
            
            # 准备总结数据并切换到总结界面
            if self.display.app:
                def show_summary_screen():
                    # 准备总结数据
                    summary_data = self._prepare_summary_data()
                    # 切换到总结界面
                    from src.display import SummaryScreen
                    summary_screen = SummaryScreen(summary_data)
                    self.display.app.push_screen(summary_screen)
                
                self.display.app.call_from_thread(show_summary_screen)
        except Exception as e:
            logger.error(f"文件处理失败: {e}", exc_info=True)
            if self.display.app:
                self.display.app.call_from_thread(self.display.app.exit)
    
    def _prepare_summary_data(self) -> dict:
        """准备总结数据"""
        duration = self.stats['end_time'] - self.stats['start_time']
        minutes = int(duration // 60)
        seconds = int(duration % 60)
        duration_str = f"{minutes}分{seconds}秒" if minutes > 0 else f"{seconds}秒"
        
        avg_speed = self.stats['success'] / duration * 60 if duration > 0 else 0
        
        # 统计输出文件
        output_files_info = []
        output_files = list(OUTPUT_DIR.glob("*.xlsx"))
        total_output_records = 0
        if output_files:
            for output_file in sorted(output_files):
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(output_file)
                    if "YS" in wb.sheetnames:
                        ws = wb["YS"]
                    else:
                        ws = wb.active
                    record_count = max(0, ws.max_row - 1)
                    total_output_records += record_count
                    output_files_info.append({
                        'name': output_file.name,
                        'records': record_count
                    })
                except Exception as e:
                    output_files_info.append({
                        'name': output_file.name,
                        'records': f"统计失败 ({e})"
                    })
        
        return {
            'total': self.stats['total'],
            'success': self.stats['success'],
            'failed': self.stats['failed'],
            'duration_str': duration_str,
            'avg_speed': avg_speed,
            'failed_files': self.stats['failed_files'],
            'output_files': output_files_info,
            'total_output_records': total_output_records,
        }
    
    def run(self):
        """运行主程序"""
        try:
            # 验证配置
            validate_config()
            
            # 检查系统依赖（如antiword）
            print("正在检查系统依赖...")
            DocumentReader.check_dependencies()
            print("系统依赖检查通过\n")
            
            # 扫描文档
            print("正在扫描文档...")
            files = self.reader.scan_documents(DATA_DIR)
            
            if not files:
                print("未找到任何文档文件！")
                return
            
            # 初始化显示（返回 App 实例）
            app = self.display.init_display()
            
            # 在后台线程中启动文件处理
            processing_thread = threading.Thread(
                target=self._process_files_in_thread,
                args=(files,),
                daemon=False
            )
            processing_thread.start()
            
            # 在主线程运行 Textual App（这会阻塞直到 App 退出）
            try:
                app.run()
            except KeyboardInterrupt:
                pass
            
            # 等待处理线程完成（总结信息会在处理完成后自动显示）
            processing_thread.join(timeout=5.0)
            
            # 清理显示（App 会在用户按 Q 键后退出）
            self.display.cleanup_display()
            
        except KeyboardInterrupt:
            self.display.cleanup_display()
            print("\n\n程序被用户中断")
            sys.exit(0)
        except Exception as e:
            self.display.cleanup_display()
            logger.error(f"程序运行失败: {e}", exc_info=True)
            print(f"\n错误: {e}")
            sys.exit(1)


def main():
    """主函数"""
    processor = DocumentProcessor()
    processor.run()


if __name__ == "__main__":
    main()

