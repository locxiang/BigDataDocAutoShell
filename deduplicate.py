"""文件去重脚本 - 基于文件hash进行去重，并清理Excel中的对应数据"""
import sys
import logging
import hashlib
from pathlib import Path
from typing import Dict, List, Tuple, Set
from collections import defaultdict
from openpyxl import load_workbook
from src.config import OUTPUT_DIR, TEMPLATE_MAPPING

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('deduplicate.log', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


class Deduplicator:
    """文件去重器"""
    
    # 支持的文档文件扩展名
    DOC_EXTENSIONS = {'.docx', '.doc', '.pdf', '.xlsx', '.xls'}
    
    # 分类文件夹到Excel文件的映射
    CATEGORY_TO_EXCEL = {
        "办会材料信息": "2办会材料信息.xlsx",
        "办文材料信息": "3办文材料信息.xlsx",
        "政策文件信息": "4政策文件信息.xlsx",
    }
    
    def __init__(self):
        """初始化去重器"""
        self.output_dir = OUTPUT_DIR
        self.stats = {
            'total_files': 0,
            'duplicate_groups': 0,
            'files_deleted': 0,
            'excel_rows_deleted': 0,
            'qa_rows_deleted': 0,
        }
    
    def calculate_file_hash(self, file_path: Path) -> str:
        """
        计算文件的MD5 hash值
        
        Args:
            file_path: 文件路径
            
        Returns:
            MD5 hash值（十六进制字符串）
        """
        hash_md5 = hashlib.md5()
        try:
            with open(file_path, 'rb') as f:
                # 分块读取，避免大文件占用过多内存
                for chunk in iter(lambda: f.read(4096), b""):
                    hash_md5.update(chunk)
            return hash_md5.hexdigest()
        except Exception as e:
            logger.error(f"计算文件hash失败: {file_path}, 错误: {e}")
            return ""
    
    def scan_files(self) -> Dict[str, List[Path]]:
        """
        扫描output目录下的所有文档文件
        
        Returns:
            字典，key为hash值，value为文件路径列表
        """
        hash_to_files = defaultdict(list)
        
        # 扫描三个分类文件夹
        for category in self.CATEGORY_TO_EXCEL.keys():
            category_dir = self.output_dir / category
            if not category_dir.exists():
                logger.info(f"分类文件夹不存在，跳过: {category_dir}")
                continue
            
            logger.info(f"扫描分类文件夹: {category_dir}")
            
            # 扫描该文件夹下的所有文档文件
            for file_path in category_dir.iterdir():
                if file_path.is_file() and file_path.suffix.lower() in self.DOC_EXTENSIONS:
                    file_hash = self.calculate_file_hash(file_path)
                    if file_hash:
                        hash_to_files[file_hash].append(file_path)
                        self.stats['total_files'] += 1
        
        return hash_to_files
    
    def find_duplicates(self, hash_to_files: Dict[str, List[Path]]) -> List[Tuple[Path, List[Path]]]:
        """
        找出重复文件组，并确定保留和删除的文件
        
        Args:
            hash_to_files: hash值到文件列表的映射
            
        Returns:
            列表，每个元素为(保留的文件路径, [要删除的文件路径列表])
        """
        duplicates = []
        
        for file_hash, file_list in hash_to_files.items():
            if len(file_list) > 1:
                # 有重复文件
                self.stats['duplicate_groups'] += 1
                
                # 按修改时间排序，保留最早的文件
                file_list.sort(key=lambda p: p.stat().st_mtime)
                
                # 第一个文件保留，其余删除
                keep_file = file_list[0]
                delete_files = file_list[1:]
                
                duplicates.append((keep_file, delete_files))
                
                logger.info(f"发现重复文件组 (hash: {file_hash[:8]}...):")
                logger.info(f"  保留: {keep_file.name} (修改时间: {self._format_time(keep_file.stat().st_mtime)})")
                for delete_file in delete_files:
                    logger.info(f"  删除: {delete_file.name} (修改时间: {self._format_time(delete_file.stat().st_mtime)})")
        
        return duplicates
    
    def _format_time(self, timestamp: float) -> str:
        """格式化时间戳为可读字符串"""
        from datetime import datetime
        return datetime.fromtimestamp(timestamp).strftime('%Y-%m-%d %H:%M:%S')
    
    def delete_file(self, file_path: Path) -> bool:
        """
        删除文件
        
        Args:
            file_path: 要删除的文件路径
            
        Returns:
            是否删除成功
        """
        try:
            file_path.unlink()
            logger.info(f"已删除文件: {file_path}")
            return True
        except Exception as e:
            logger.error(f"删除文件失败: {file_path}, 错误: {e}")
            return False
    
    def get_category_from_path(self, file_path: Path) -> str:
        """
        从文件路径中提取分类名称
        
        Args:
            file_path: 文件路径
            
        Returns:
            分类名称，如果无法确定则返回None
        """
        # 文件路径格式: output/分类名称/文件名
        parent_dir = file_path.parent.name
        if parent_dir in self.CATEGORY_TO_EXCEL:
            return parent_dir
        return None
    
    def delete_excel_row_by_filename(self, excel_file: Path, filename_without_ext: str) -> int:
        """
        根据文件名删除Excel中对应的行
        
        Args:
            excel_file: Excel文件路径
            filename_without_ext: 文件名（不含扩展名）
            
        Returns:
            删除的行数
        """
        if not excel_file.exists():
            logger.warning(f"Excel文件不存在: {excel_file}")
            return 0
        
        try:
            wb = load_workbook(excel_file)
            if "YS" not in wb.sheetnames:
                logger.warning(f"Excel文件中没有YS Sheet: {excel_file}")
                wb.close()
                return 0
            
            ws = wb["YS"]
            
            # 查找PolicyFileName列
            header_row = 1
            policy_filename_col = None
            
            # 查找表头中的PolicyFileName列
            for col_idx, cell in enumerate(ws[header_row], start=1):
                header_value = cell.value if cell.value else ""
                # 提取键名（去掉中文注释）
                key = self._extract_key_from_header(str(header_value))
                if key == "PolicyFileName":
                    policy_filename_col = col_idx
                    break
            
            if policy_filename_col is None:
                logger.warning(f"Excel文件中未找到PolicyFileName列: {excel_file}")
                wb.close()
                return 0
            
            # 查找匹配的行（从后往前删除，避免索引变化）
            rows_to_delete = []
            for row_idx in range(ws.max_row, header_row, -1):
                cell_value = ws.cell(row_idx, policy_filename_col).value
                if cell_value:
                    # 比较文件名（不含扩展名）
                    cell_filename = str(cell_value).strip()
                    if cell_filename == filename_without_ext:
                        rows_to_delete.append(row_idx)
            
            # 删除匹配的行
            deleted_count = 0
            for row_idx in rows_to_delete:
                ws.delete_rows(row_idx)
                deleted_count += 1
            
            if deleted_count > 0:
                wb.save(excel_file)
                logger.info(f"从 {excel_file.name} 中删除了 {deleted_count} 行数据 (文件名: {filename_without_ext})")
            
            wb.close()
            return deleted_count
            
        except Exception as e:
            logger.error(f"删除Excel行失败: {excel_file}, 错误: {e}")
            return 0
    
    def delete_qa_rows_by_policy_id(self, qa_excel_file: Path, policy_ids: Set[int]) -> int:
        """
        根据政策ID删除问答对Excel中对应的行
        
        Args:
            qa_excel_file: 问答对Excel文件路径
            policy_ids: 政策ID集合
            
        Returns:
            删除的行数
        """
        if not qa_excel_file.exists():
            logger.warning(f"问答对Excel文件不存在: {qa_excel_file}")
            return 0
        
        if not policy_ids:
            return 0
        
        try:
            wb = load_workbook(qa_excel_file)
            if "YS" not in wb.sheetnames:
                logger.warning(f"问答对Excel文件中没有YS Sheet: {qa_excel_file}")
                wb.close()
                return 0
            
            ws = wb["YS"]
            
            # 查找ID列
            header_row = 1
            id_col = None
            
            # 查找表头中的ID列
            for col_idx, cell in enumerate(ws[header_row], start=1):
                header_value = cell.value if cell.value else ""
                key = self._extract_key_from_header(str(header_value))
                if key == "ID":
                    id_col = col_idx
                    break
            
            if id_col is None:
                logger.warning(f"问答对Excel文件中未找到ID列: {qa_excel_file}")
                wb.close()
                return 0
            
            # 查找匹配的行（从后往前删除）
            rows_to_delete = []
            for row_idx in range(ws.max_row, header_row, -1):
                cell_value = ws.cell(row_idx, id_col).value
                if cell_value:
                    try:
                        policy_id = int(cell_value)
                        if policy_id in policy_ids:
                            rows_to_delete.append(row_idx)
                    except (ValueError, TypeError):
                        continue
            
            # 删除匹配的行
            deleted_count = 0
            for row_idx in rows_to_delete:
                ws.delete_rows(row_idx)
                deleted_count += 1
            
            if deleted_count > 0:
                wb.save(qa_excel_file)
                logger.info(f"从 {qa_excel_file.name} 中删除了 {deleted_count} 行问答对数据 (政策ID: {policy_ids})")
            
            wb.close()
            return deleted_count
            
        except Exception as e:
            logger.error(f"删除问答对行失败: {qa_excel_file}, 错误: {e}")
            return 0
    
    def get_policy_ids_by_filename(self, excel_file: Path, filename_without_ext: str) -> Set[int]:
        """
        根据文件名获取政策ID列表
        
        Args:
            excel_file: Excel文件路径
            filename_without_ext: 文件名（不含扩展名）
            
        Returns:
            政策ID集合
        """
        policy_ids = set()
        
        if not excel_file.exists():
            return policy_ids
        
        try:
            wb = load_workbook(excel_file)
            if "YS" not in wb.sheetnames:
                wb.close()
                return policy_ids
            
            ws = wb["YS"]
            
            # 查找PolicyFileName列和ID列
            header_row = 1
            policy_filename_col = None
            id_col = None
            
            for col_idx, cell in enumerate(ws[header_row], start=1):
                header_value = cell.value if cell.value else ""
                key = self._extract_key_from_header(str(header_value))
                if key == "PolicyFileName":
                    policy_filename_col = col_idx
                elif key == "ID":
                    id_col = col_idx
            
            if policy_filename_col is None or id_col is None:
                wb.close()
                return policy_ids
            
            # 查找匹配的行，提取ID
            for row_idx in range(2, ws.max_row + 1):
                filename_value = ws.cell(row_idx, policy_filename_col).value
                if filename_value and str(filename_value).strip() == filename_without_ext:
                    id_value = ws.cell(row_idx, id_col).value
                    if id_value:
                        try:
                            policy_id = int(id_value)
                            policy_ids.add(policy_id)
                        except (ValueError, TypeError):
                            continue
            
            wb.close()
            return policy_ids
            
        except Exception as e:
            logger.error(f"获取政策ID失败: {excel_file}, 错误: {e}")
            return policy_ids
    
    @staticmethod
    def _extract_key_from_header(header: str) -> str:
        """
        从表头中提取英文键名
        
        Args:
            header: 表头字符串，例如 "ID(序号)" 或 "PolicyCategory(文件类别)"
            
        Returns:
            提取的键名，例如 "ID" 或 "PolicyCategory"
        """
        if not header:
            return ""
        
        # 如果包含括号，提取括号前的部分
        if "(" in header:
            key = header.split("(")[0].strip()
            return key
        
        # 如果没有括号，直接返回
        return header.strip()
    
    def process_duplicates(self, duplicates: List[Tuple[Path, List[Path]]]):
        """
        处理重复文件：删除文件并清理Excel数据
        
        Args:
            duplicates: 重复文件列表
        """
        # 用于收集政策文件的ID（用于删除问答对）
        policy_ids_to_delete = set()
        
        for keep_file, delete_files in duplicates:
            category = self.get_category_from_path(keep_file)
            if not category:
                logger.warning(f"无法确定文件分类，跳过: {keep_file}")
                continue
            
            excel_file = self.output_dir / self.CATEGORY_TO_EXCEL[category]
            
            # 处理每个要删除的文件
            for delete_file in delete_files:
                # 1. 提取文件名（不含扩展名）
                filename_without_ext = delete_file.stem
                
                # 2. 如果是政策文件，先收集ID（在删除Excel行之前）
                if category == "政策文件信息":
                    policy_ids = self.get_policy_ids_by_filename(excel_file, filename_without_ext)
                    policy_ids_to_delete.update(policy_ids)
                
                # 3. 删除Excel中对应的行
                deleted_rows = self.delete_excel_row_by_filename(excel_file, filename_without_ext)
                self.stats['excel_rows_deleted'] += deleted_rows
                
                # 4. 删除文件
                if self.delete_file(delete_file):
                    self.stats['files_deleted'] += 1
        
        # 5. 删除政策问答对中关联的记录
        if policy_ids_to_delete:
            qa_excel_file = self.output_dir / TEMPLATE_MAPPING["政策问答对"]
            deleted_qa_rows = self.delete_qa_rows_by_policy_id(qa_excel_file, policy_ids_to_delete)
            self.stats['qa_rows_deleted'] = deleted_qa_rows
    
    def print_summary(self):
        """打印统计信息"""
        print("\n" + "=" * 60)
        print("去重完成统计")
        print("=" * 60)
        print(f"扫描文件总数: {self.stats['total_files']}")
        print(f"发现重复文件组: {self.stats['duplicate_groups']}")
        print(f"删除文件数量: {self.stats['files_deleted']}")
        print(f"删除Excel行数: {self.stats['excel_rows_deleted']}")
        print(f"删除问答对行数: {self.stats['qa_rows_deleted']}")
        print("=" * 60)
    
    def run(self):
        """运行去重流程"""
        try:
            print("=" * 60)
            print("文件去重脚本")
            print("=" * 60)
            print(f"输出目录: {self.output_dir}")
            print()
            
            # 1. 扫描文件
            print("步骤1: 扫描文件...")
            hash_to_files = self.scan_files()
            print(f"扫描完成，共找到 {self.stats['total_files']} 个文件")
            print()
            
            if self.stats['total_files'] == 0:
                print("未找到任何文件，退出")
                return
            
            # 2. 查找重复文件
            print("步骤2: 查找重复文件...")
            duplicates = self.find_duplicates(hash_to_files)
            print(f"查找完成，共发现 {self.stats['duplicate_groups']} 组重复文件")
            print()
            
            if self.stats['duplicate_groups'] == 0:
                print("未发现重复文件，退出")
                return
            
            # 3. 确认操作
            total_files_to_delete = sum(len(delete_files) for _, delete_files in duplicates)
            print(f"将删除 {total_files_to_delete} 个重复文件")
            print("保留策略: 保留最早的文件（按修改时间）")
            response = input("\n是否继续？(y/n): ").strip().lower()
            
            if response != 'y':
                print("操作已取消")
                return
            
            print()
            
            # 4. 处理重复文件
            print("步骤3: 删除重复文件并清理Excel数据...")
            self.process_duplicates(duplicates)
            print()
            
            # 5. 打印统计信息
            self.print_summary()
            
        except KeyboardInterrupt:
            print("\n\n程序被用户中断")
            sys.exit(0)
        except Exception as e:
            logger.error(f"程序运行失败: {e}", exc_info=True)
            print(f"\n错误: {e}")
            sys.exit(1)


def main():
    """主函数"""
    deduplicator = Deduplicator()
    deduplicator.run()


if __name__ == "__main__":
    main()

