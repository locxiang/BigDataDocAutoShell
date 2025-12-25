"""政策问答对生成模块 - 基于政策文件生成问答对"""
import logging
import json
import re
import threading
from pathlib import Path
from typing import List, Dict, Any, Optional
from openai import OpenAI
from openpyxl import load_workbook, Workbook
from src.config import (
    OPENAI_API_KEY, OPENAI_BASE_URL, MODEL_NAME, 
    MAX_RETRIES, REQUEST_TIMEOUT, OUTPUT_DIR, TEMPLATE_DIR, TEMPLATE_MAPPING
)

logger = logging.getLogger(__name__)


class QAGenerator:
    """政策问答对生成器"""
    
    # 类级别的文件锁，保护Excel文件写入
    _file_lock = threading.Lock()
    
    # 生成问答对的Prompt模板
    QA_GENERATION_PROMPT = """你是一个专业的政策解读专家。请根据以下政策文件信息，生成10个问答对。

政策文件信息：
- 政策名称：{policy_name}
- 发文文号：{document_number}
- 发布单位：{issuing_authority}
- 成文日期：{effective_date}
- 施行日期：{implementation_date}
- 有效期至：{valid_until}
- 责任部门：{responsible_department}
- 协同部门：{collaborative_department}
- 适用对象：{applicable_object}
- 涉及领域：{fields}
- 政策类别：{policy_category}
- 备注（政策标题）：{remarks}

请生成10个问答对，要求：
1. **问题设计**：
   - 问题要贴近老百姓的实际需求，从普通民众的角度提问
   - 问题要具体、实用，避免过于抽象
   - 可以包括：政策适用范围、申请条件、办理流程、所需材料、办理时限、咨询方式、政策有效期、政策对象、政策内容等
   - 问题表述要通俗易懂，符合老百姓的日常用语习惯
   - 问题要多样化，涵盖政策的不同方面

2. **答案设计**：
   - 答案要准确、完整，基于政策文件信息
   - 语言要符合政府文件的用词用语习惯，使用规范的政府术语
   - 但表述要清晰明了，让老百姓能够理解
   - 可以适当解释专业术语，但不要过于口语化
   - 如果涉及具体流程或要求，要详细说明
   - 如果信息不完整，可以基于政策名称和类别进行合理推断，但要标注"具体以政策文件为准"

3. **风格要求**：
   - 问答对要体现政府文件的权威性和规范性
   - 同时要兼顾可读性，让普通民众能够理解
   - 避免使用过于生硬的官方套话，但也不能过于随意
   - 保持专业、准确、易懂的平衡
   - 使用规范的政府用语，如"根据...规定"、"按照...要求"等

4. **内容要求**：
   - 必须生成恰好10个问答对，不能多也不能少
   - 每个问答对都要有实际意义，不能重复
   - 问答对要覆盖政策的主要方面

请严格按照以下JSON格式返回，不要添加任何其他文字：
{{
  "qa_pairs": [
    {{
      "question": "问题1",
      "answer": "答案1"
    }},
    {{
      "question": "问题2",
      "answer": "答案2"
    }},
    {{
      "question": "问题3",
      "answer": "答案3"
    }},
    {{
      "question": "问题4",
      "answer": "答案4"
    }},
    {{
      "question": "问题5",
      "answer": "答案5"
    }},
    {{
      "question": "问题6",
      "answer": "答案6"
    }},
    {{
      "question": "问题7",
      "answer": "答案7"
    }},
    {{
      "question": "问题8",
      "answer": "答案8"
    }},
    {{
      "question": "问题9",
      "answer": "答案9"
    }},
    {{
      "question": "问题10",
      "answer": "答案10"
    }}
  ]
}}"""
    
    def __init__(self):
        """初始化生成器"""
        if not OPENAI_API_KEY:
            raise ValueError("OPENAI_API_KEY 未设置")
        
        self.client = OpenAI(
            api_key=OPENAI_API_KEY,
            base_url=OPENAI_BASE_URL,
            timeout=REQUEST_TIMEOUT,
        )
        self.model = MODEL_NAME
    
    def read_policy_file(self, excel_path: Path) -> List[Dict[str, Any]]:
        """
        读取政策文件Excel
        
        Args:
            excel_path: Excel文件路径
            
        Returns:
            政策文件列表
        """
        try:
            wb = load_workbook(excel_path)
            if "YS" in wb.sheetnames:
                ws = wb["YS"]
            else:
                ws = wb.active
            
            # 读取表头
            headers = []
            if ws.max_row > 0:
                for cell in ws[1]:
                    header_value = cell.value if cell.value else ""
                    headers.append(header_value)
            
            # 提取键名（去掉中文注释）
            header_keys = []
            for header in headers:
                key = self._extract_key_from_header(header)
                header_keys.append(key)
            
            # 读取数据行
            policies = []
            for row_idx in range(2, ws.max_row + 1):
                policy = {}
                for col_idx, key in enumerate(header_keys, start=1):
                    cell_value = ws.cell(row_idx, col_idx).value
                    if cell_value is not None:
                        policy[key] = str(cell_value).strip()
                    else:
                        policy[key] = ""
                
                # 只添加有ID的政策（跳过空行）
                if policy.get("ID") and policy.get("ID") != "":
                    policies.append(policy)
            
            logger.info(f"从 {excel_path.name} 读取到 {len(policies)} 条政策记录")
            return policies
            
        except Exception as e:
            logger.error(f"读取政策文件失败: {e}")
            raise
    
    def generate_qa_pairs(self, policy: Dict[str, Any]) -> List[Dict[str, str]]:
        """
        为单个政策生成10个问答对
        
        Args:
            policy: 政策信息字典
            
        Returns:
            问答对列表，每个元素包含question和answer
        """
        # 准备Prompt参数
        prompt_params = {
            "policy_name": policy.get("Remarks", policy.get("PolicyFileName", "")),
            "document_number": policy.get("DocumentNumber", ""),
            "issuing_authority": policy.get("IssuingAuthority", ""),
            "effective_date": policy.get("EffectiveDate", ""),
            "implementation_date": policy.get("ImplementationDate", ""),
            "valid_until": policy.get("ValidUntil", ""),
            "responsible_department": policy.get("ResponsibleDepartment", ""),
            "collaborative_department": policy.get("CollaborativeDepartment", ""),
            "applicable_object": policy.get("ApplicableObject", ""),
            "fields": policy.get("Fields", ""),
            "policy_category": policy.get("PolicyCategory", ""),
            "remarks": policy.get("Remarks", ""),
        }
        
        # 格式化Prompt
        prompt = self.QA_GENERATION_PROMPT.format(**prompt_params)
        
        for attempt in range(MAX_RETRIES):
            try:
                response = self.client.chat.completions.create(
                    model=self.model,
                    messages=[
                        {"role": "user", "content": prompt}
                    ],
                    temperature=0.7,  # 稍微提高温度以增加多样性
                )
                
                result_text = response.choices[0].message.content.strip()
                
                # 解析JSON
                qa_data = self._parse_json(result_text)
                
                if qa_data and "qa_pairs" in qa_data:
                    qa_pairs = qa_data["qa_pairs"]
                    if isinstance(qa_pairs, list) and len(qa_pairs) > 0:
                        logger.info(f"成功生成 {len(qa_pairs)} 个问答对")
                        return qa_pairs
                    else:
                        logger.warning(f"问答对列表为空: {policy.get('ID', 'unknown')}")
                        if attempt < MAX_RETRIES - 1:
                            continue
                        return []
                else:
                    logger.warning(f"JSON解析失败: {policy.get('ID', 'unknown')}, 返回: {result_text[:200]}")
                    if attempt < MAX_RETRIES - 1:
                        continue
                    return []
                    
            except Exception as e:
                logger.error(f"生成问答对失败 (尝试 {attempt + 1}/{MAX_RETRIES}): {policy.get('ID', 'unknown')}, 错误: {e}")
                if attempt < MAX_RETRIES - 1:
                    continue
                return []
        
        return []
    
    @staticmethod
    def _parse_json(text: str) -> Optional[Dict[str, Any]]:
        """
        从文本中解析JSON
        
        Args:
            text: 包含JSON的文本
            
        Returns:
            解析后的字典，失败返回None
        """
        # 尝试直接解析
        try:
            return json.loads(text)
        except json.JSONDecodeError:
            pass
        
        # 尝试提取JSON代码块
        json_match = re.search(r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\}', text, re.DOTALL)
        if json_match:
            try:
                return json.loads(json_match.group())
            except json.JSONDecodeError:
                pass
        
        # 尝试提取markdown代码块中的JSON
        code_block_match = re.search(r'```(?:json)?\s*(\{.*?\})\s*```', text, re.DOTALL)
        if code_block_match:
            try:
                return json.loads(code_block_match.group(1))
            except json.JSONDecodeError:
                pass
        
        return None
    
    @staticmethod
    def _extract_key_from_header(header: str) -> str:
        """
        从表头中提取英文键名
        
        Args:
            header: 表头字符串，例如 "ID(序号)" 或 "PolicyCategory(文件类别)"
            
        Returns:
            提取的键名，例如 "ID" 或 "PolicyCategory"
        """
        if not header:
            return ""
        
        # 如果包含括号，提取括号前的部分
        if "(" in header:
            key = header.split("(")[0].strip()
            return key
        
        # 如果没有括号，直接返回
        return header.strip()
    
    def save_qa_pairs(self, qa_pairs: List[Dict[str, Any]], policy: Dict[str, Any]) -> bool:
        """
        保存问答对到Excel（线程安全）
        
        Args:
            qa_pairs: 问答对列表
            policy: 政策信息字典
            
        Returns:
            是否保存成功
        """
        # 使用类级别的锁保护Excel文件写入操作
        with self._file_lock:
            try:
                # 获取输出文件路径
                template_name = TEMPLATE_MAPPING.get("政策问答对")
                if not template_name:
                    raise ValueError("未找到政策问答对模板文件")
                
                output_file = OUTPUT_DIR / template_name
                
                # 如果输出文件不存在，从模板复制
                if not output_file.exists():
                    template_file = TEMPLATE_DIR / template_name
                    if template_file.exists():
                        import shutil
                        shutil.copy2(template_file, output_file)
                        logger.info(f"从模板创建输出文件: {output_file}")
                    else:
                        # 创建新的Excel文件
                        wb = Workbook()
                        ws = wb.active
                        ws.title = "YS"
                        # 添加表头
                        headers = ["XH", "ID", "SourceDepartment", "Problem", "Answer", 
                                  "PolicyName", "DocumentNumber", "Fields", "Remarks"]
                        ws.append(headers)
                        wb.save(output_file)
                        logger.info(f"创建新的输出文件: {output_file}")
                
                # 加载工作簿
                wb = load_workbook(output_file)
                if "YS" in wb.sheetnames:
                    ws = wb["YS"]
                else:
                    ws = wb.active
                    if ws.title != "YS":
                        ws.title = "YS"
                
                # 获取表头
                headers = []
                header_keys = []
                if ws.max_row > 0:
                    for cell in ws[1]:
                        header_value = cell.value if cell.value else ""
                        headers.append(header_value)
                        key = self._extract_key_from_header(header_value)
                        header_keys.append(key)
                
                # 如果没有表头，创建表头
                if not headers or not any(headers):
                    headers = ["XH", "ID", "SourceDepartment", "Problem", "Answer", 
                              "PolicyName", "DocumentNumber", "Fields", "Remarks"]
                    header_keys = headers.copy()
                    ws.append(headers)
                
                # 获取当前最大序号
                max_xh = 0
                if "XH" in header_keys:
                    xh_col = header_keys.index("XH") + 1
                    for row in range(2, ws.max_row + 1):
                        cell_value = ws.cell(row, xh_col).value
                        if isinstance(cell_value, (int, float)):
                            max_xh = max(max_xh, int(cell_value))
                        elif isinstance(cell_value, str) and cell_value.strip().isdigit():
                            max_xh = max(max_xh, int(cell_value.strip()))
                
                # 准备政策信息
                policy_id = policy.get("ID", "")
                source_department = policy.get("ResponsibleDepartment", policy.get("IssuingAuthority", ""))
                policy_name = policy.get("Remarks", policy.get("PolicyFileName", ""))
                document_number = policy.get("DocumentNumber", "")
                fields = policy.get("Fields", "")
                remarks = policy.get("Remarks", "")
                
                # 保存每个问答对
                for idx, qa_pair in enumerate(qa_pairs):
                    row_data = []
                    for key in header_keys:
                        if key == "XH":
                            value = max_xh + idx + 1
                        elif key == "ID":
                            value = policy_id
                        elif key == "SourceDepartment":
                            value = source_department
                        elif key == "Problem":
                            value = qa_pair.get("question", "")
                        elif key == "Answer":
                            value = qa_pair.get("answer", "")
                        elif key == "PolicyName":
                            value = policy_name
                        elif key == "DocumentNumber":
                            value = document_number
                        elif key == "Fields":
                            value = fields
                        elif key == "Remarks":
                            value = remarks
                        else:
                            value = ""
                        
                        row_data.append(value)
                    
                ws.append(row_data)
                
                # 保存文件
                wb.save(output_file)
                logger.info(f"成功保存 {len(qa_pairs)} 个问答对到: {output_file.name}")
                return True
                
            except Exception as e:
                logger.error(f"保存问答对失败: {e}")
                return False

