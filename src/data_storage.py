"""数据存储模块 - 将提取的信息写入Excel"""
import logging
import threading
from pathlib import Path
from typing import Dict, Any, Optional
from openpyxl import load_workbook, Workbook
from src.config import TEMPLATE_DIR, OUTPUT_DIR, TEMPLATE_MAPPING

logger = logging.getLogger(__name__)


class DataStorage:
    """数据存储器"""
    
    # 类级别的锁字典，为每个文档类型的Excel文件提供独立的锁
    _file_locks = {}
    _locks_lock = threading.Lock()  # 保护锁字典本身的锁
    
    def __init__(self):
        """初始化存储器"""
        self.output_dir = OUTPUT_DIR
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    @classmethod
    def _get_file_lock(cls, doc_type: str) -> threading.Lock:
        """
        获取指定文档类型的文件锁
        
        Args:
            doc_type: 文档类型
            
        Returns:
            对应的线程锁
        """
        with cls._locks_lock:
            if doc_type not in cls._file_locks:
                cls._file_locks[doc_type] = threading.Lock()
            return cls._file_locks[doc_type]
    
    def get_output_file(self, doc_type: str) -> Path:
        """
        获取输出Excel文件路径
        
        Args:
            doc_type: 文档类型
            
        Returns:
            输出文件路径
        """
        template_name = TEMPLATE_MAPPING.get(doc_type)
        if not template_name:
            raise ValueError(f"未知的文档类型: {doc_type}")
        
        output_file = self.output_dir / template_name
        
        # 如果输出文件不存在，从模板复制
        if not output_file.exists():
            template_file = TEMPLATE_DIR / template_name
            if template_file.exists():
                import shutil
                shutil.copy2(template_file, output_file)
                logger.info(f"从模板创建输出文件: {output_file}")
            else:
                # 创建新的Excel文件
                wb = Workbook()
                ws = wb.active
                ws.title = "YS"  # 使用YS作为Sheet名称
                wb.save(output_file)
                logger.info(f"创建新的输出文件: {output_file}")
        
        return output_file
    
    def save_data(self, data: Dict[str, Any], doc_type: str) -> bool:
        """
        保存数据到Excel（线程安全）
        
        Args:
            data: 提取的数据字典
            doc_type: 文档类型
            
        Returns:
            是否保存成功
        """
        # 获取该文档类型的文件锁
        file_lock = self._get_file_lock(doc_type)
        
        # 使用锁保护Excel文件写入操作
        with file_lock:
            try:
                output_file = self.get_output_file(doc_type)
                wb = load_workbook(output_file)
                # 所有类型（办文、办会、政策文件）都使用YS Sheet
                if "YS" in wb.sheetnames:
                    ws = wb["YS"]
                else:
                    # 如果没有YS Sheet，创建它
                    ws = wb.active
                    if ws.title != "YS":
                        ws.title = "YS"
                
                # 获取表头（YS Sheet使用标准格式，第一行就是表头）
                headers = []
                header_keys = []  # 用于匹配数据的键名
                
                if ws.max_row > 0:
                    # 标准格式：第一行就是表头
                    for cell in ws[1]:
                        header_value = cell.value if cell.value else ""
                        headers.append(header_value)
                        # 从表头中提取英文键名（去掉中文注释部分）
                        # 例如 "ID(序号)" -> "ID", "PolicyCategory(文件类别)" -> "PolicyCategory"
                        key = self._extract_key_from_header(header_value)
                        header_keys.append(key)
                
                # 如果没有表头，根据文档类型创建
                if not headers or not any(headers):
                    headers = self._get_headers_for_type(doc_type)
                    header_keys = headers.copy()
                    # 标准格式：直接添加表头
                    ws.append(headers)
                
                # 准备数据行
                row_data = []
                for key in header_keys:
                    # 如果是op列，统一设置为"insert"
                    if key.lower() == "op":
                        value = "insert"
                    else:
                        value = data.get(key, "")
                        # 处理空值
                        if value is None:
                            value = ""
                        # 处理列表类型（如Fields字段可能是列表）
                        elif isinstance(value, list):
                            # 将列表转换为字符串，用逗号分隔
                            value = "、".join(str(item) for item in value if item)
                        # 确保值是字符串或数字类型（openpyxl支持的类型）
                        elif not isinstance(value, (str, int, float, bool)):
                            value = str(value)
                    row_data.append(value)
                
                # YS Sheet使用标准格式：追加数据行
                # 添加ID（序号）
                if "ID" in header_keys:
                    # ID列通常是第一列
                    id_col = header_keys.index("ID")
                    
                    # 获取当前最大ID
                    max_id = 0
                    for row in range(2, ws.max_row + 1):
                        cell_value = ws.cell(row, id_col + 1).value
                        if isinstance(cell_value, (int, float)):
                            max_id = max(max_id, int(cell_value))
                    
                    row_data[id_col] = max_id + 1
                
                # 追加数据行
                ws.append(row_data)
                
                # 保存文件
                wb.save(output_file)
                logger.info(f"数据已保存到: {output_file}")
                return True
                
            except Exception as e:
                logger.error(f"保存数据失败: {e}")
                return False
    
    @staticmethod
    def _get_headers_for_type(doc_type: str) -> list[str]:
        """
        根据文档类型获取表头
        
        Args:
            doc_type: 文档类型
            
        Returns:
            表头列表
        """
        headers_map = {
            "办会材料信息": [
                "ID", "PolicyCategory", "PolicyFileName", "IssuingAuthority",
                "EffectiveDate", "Position", "Topic", "Refrence", "Remarks"
            ],
            "办文材料信息": [
                "ID", "PolicyCategory", "PolicyFileName", "IssuingAuthority",
                "EffectiveDate", "Position", "ObjectOriented", "Topic",
                "Refrence", "Language", "Remarks"
            ],
            "政策文件信息": [
                "ID", "PolicyCategory", "PolicyFileName", "DocumentNumber",
                "IssuingAuthority", "EffectiveDate", "ImplementationDate",
                "ValidUntil", "ResponsibleDepartment", "CollaborativeDepartment",
                "ApplicableObject", "Fields", "Remarks"
            ],
        }
        
        return headers_map.get(doc_type, [])
    
    @staticmethod
    def _extract_key_from_header(header: str) -> str:
        """
        从表头中提取英文键名
        
        Args:
            header: 表头字符串，例如 "ID(序号)" 或 "PolicyCategory(文件类别)"
            
        Returns:
            提取的键名，例如 "ID" 或 "PolicyCategory"
        """
        if not header:
            return ""
        
        # 如果包含括号，提取括号前的部分
        if "(" in header:
            key = header.split("(")[0].strip()
            return key
        
        # 如果没有括号，直接返回
        return header.strip()

