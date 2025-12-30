"""文件名规范化脚本 - 处理文件名中的空格和括号问题，并同步更新Excel"""
import sys
import logging
import re
import hashlib
from pathlib import Path
from typing import List, Tuple, Dict
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook
from src.config import OUTPUT_DIR, TEMPLATE_MAPPING

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('normalize_filenames.log', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


class FilenameNormalizer:
    """文件名规范化器"""
    
    # 支持的文档文件扩展名
    DOC_EXTENSIONS = {'.docx', '.doc', '.pdf', '.xlsx', '.xls'}
    
    # 分类文件夹到Excel文件的映射
    CATEGORY_TO_EXCEL = {
        "办会材料信息": "2办会材料信息.xlsx",
        "办文材料信息": "3办文材料信息.xlsx",
        "政策文件信息": "4政策文件信息.xlsx",
    }
    
    def __init__(self):
        """初始化规范化器"""
        self.output_dir = OUTPUT_DIR
        self.stats = {
            'total_files': 0,
            'files_renamed': 0,
            'excel_rows_updated': 0,
            'errors': 0,
        }
    
    def normalize_filename(self, filename: str) -> str:
        """
        规范化文件名
        
        规则：
        1. 空格、Tab等空白字符 -> 下划线
        2. 英文圆括号 () -> 中文圆括号 （）
        
        Args:
            filename: 原始文件名（不含路径，但包含扩展名）
            
        Returns:
            规范化后的文件名
        """
        # 分离文件名和扩展名
        path_obj = Path(filename)
        stem = path_obj.stem  # 不含扩展名的文件名
        suffix = path_obj.suffix  # 扩展名
        
        # 1. 替换空白字符（空格、Tab、换行等）为下划线
        # 使用正则表达式匹配所有空白字符
        normalized_stem = re.sub(r'\s+', '_', stem)
        
        # 2. 替换英文圆括号为中文圆括号
        normalized_stem = normalized_stem.replace('(', '（').replace(')', '）')
        
        # 组合文件名和扩展名
        normalized_filename = normalized_stem + suffix
        
        return normalized_filename
    
    def needs_normalization(self, filename: str) -> bool:
        """
        检查文件名是否需要规范化（快速检查，不计算规范化结果）
        
        Args:
            filename: 文件名
            
        Returns:
            是否需要规范化
        """
        # 快速检查：是否存在空白字符或英文括号
        path_obj = Path(filename)
        stem = path_obj.stem
        # 检查是否有空白字符
        if re.search(r'\s', stem):
            return True
        # 检查是否有英文括号
        if '(' in stem or ')' in stem:
            return True
        return False
    
    def scan_files(self) -> Dict[str, List[Tuple[Path, str]]]:
        """
        扫描output目录下的所有文档文件，按分类分组
        
        Returns:
            字典，key为分类名称，value为该分类下需要处理的文件列表
        """
        files_by_category = defaultdict(list)
        
        # 扫描三个分类文件夹
        for category in self.CATEGORY_TO_EXCEL.keys():
            category_dir = self.output_dir / category
            if not category_dir.exists():
                logger.info(f"分类文件夹不存在，跳过: {category_dir}")
                continue
            
            logger.info(f"扫描分类文件夹: {category_dir}")
            
            # 扫描该文件夹下的所有文档文件
            for file_path in category_dir.iterdir():
                if file_path.is_file() and file_path.suffix.lower() in self.DOC_EXTENSIONS:
                    self.stats['total_files'] += 1
                    
                    # 检查是否需要规范化
                    if self.needs_normalization(file_path.name):
                        files_by_category[category].append((file_path, category))
        
        return files_by_category
    
    def rename_file(self, file_path: Path, new_filename: str) -> bool:
        """
        重命名文件
        
        Args:
            file_path: 原文件路径
            new_filename: 新文件名
            
        Returns:
            是否重命名成功
        """
        try:
            old_filename = file_path.name
            new_path = file_path.parent / new_filename
            
            # 检查新文件名是否已存在
            if new_path.exists() and new_path != file_path:
                logger.warning(f"目标文件已存在，跳过重命名: {new_path}")
                return False
            
            file_path.rename(new_path)
            logger.info(f"文件重命名成功: {old_filename} -> {new_filename}")
            return True
        except Exception as e:
            logger.error(f"文件重命名失败: {file_path}, 错误: {e}")
            return False
    
    @staticmethod
    def _extract_key_from_header(header: str) -> str:
        """
        从表头中提取英文键名
        
        Args:
            header: 表头字符串，例如 "ID(序号)" 或 "PolicyCategory(文件类别)"
            
        Returns:
            提取的键名，例如 "ID" 或 "PolicyCategory"
        """
        if not header:
            return ""
        
        # 如果包含括号，提取括号前的部分
        if "(" in header:
            key = header.split("(")[0].strip()
            return key
        
        # 如果没有括号，直接返回
        return header.strip()
    
    def batch_update_excel_filenames(self, excel_file: Path, filename_mappings: Dict[str, str]) -> int:
        """
        批量更新Excel中PolicyFileName字段的值（优化版本：只打开一次Excel文件）
        
        Args:
            excel_file: Excel文件路径
            filename_mappings: 文件名映射字典，key为旧文件名（不含扩展名），value为新文件名（不含扩展名）
            
        Returns:
            更新的行数
        """
        if not excel_file.exists():
            logger.warning(f"Excel文件不存在: {excel_file}")
            return 0
        
        if not filename_mappings:
            return 0
        
        try:
            logger.info(f"打开Excel文件进行批量更新: {excel_file.name}，共 {len(filename_mappings)} 个文件名需要更新")
            wb = load_workbook(excel_file)
            if "YS" not in wb.sheetnames:
                logger.warning(f"Excel文件中没有YS Sheet: {excel_file}")
                wb.close()
                return 0
            
            ws = wb["YS"]
            
            # 查找PolicyFileName列
            header_row = 1
            policy_filename_col = None
            
            # 查找表头中的PolicyFileName列
            for col_idx, cell in enumerate(ws[header_row], start=1):
                header_value = cell.value if cell.value else ""
                key = self._extract_key_from_header(str(header_value))
                if key == "PolicyFileName":
                    policy_filename_col = col_idx
                    break
            
            if policy_filename_col is None:
                logger.warning(f"Excel文件中未找到PolicyFileName列: {excel_file}")
                wb.close()
                return 0
            
            # 批量查找匹配的行并更新（使用字典快速查找）
            updated_count = 0
            for row_idx in range(2, ws.max_row + 1):
                cell_value = ws.cell(row_idx, policy_filename_col).value
                if cell_value:
                    # 比较文件名（不含扩展名）
                    cell_filename = str(cell_value).strip()
                    if cell_filename in filename_mappings:
                        # 更新为新文件名
                        new_filename = filename_mappings[cell_filename]
                        ws.cell(row_idx, policy_filename_col).value = new_filename
                        updated_count += 1
                        logger.debug(f"更新Excel行 {row_idx}: {cell_filename} -> {new_filename}")
            
            if updated_count > 0:
                wb.save(excel_file)
                logger.info(f"从 {excel_file.name} 中批量更新了 {updated_count} 行数据")
            
            wb.close()
            return updated_count
            
        except Exception as e:
            logger.error(f"批量更新Excel失败: {excel_file}, 错误: {e}")
            return 0
    
    def check_and_resolve_duplicates(self, files_list: List[Tuple[Path, str]], category_dir: Path) -> Dict[Path, str]:
        """
        检查并解决重复文件名问题（优化版本：排序后只检查相邻文件）
        
        Args:
            files_list: 需要处理的文件列表
            category_dir: 分类目录路径
            
        Returns:
            文件名映射字典，key为文件路径，value为最终确定的新文件名
        """
        # 1. 计算所有新文件名并创建(文件名, 文件路径)元组列表
        file_info_list = []
        for file_path, _ in files_list:
            old_filename = file_path.name
            new_filename = self.normalize_filename(old_filename)
            file_info_list.append((new_filename, file_path))
        
        # 2. 按文件名排序（排序后重复的文件名会聚集在一起）
        file_info_list.sort(key=lambda x: x[0])
        
        # 3. 获取目录中已存在的文件名（用于检查冲突，排除当前正在处理的文件）
        existing_files = set()
        processing_filenames = {file_path.name for _, file_path in file_info_list}
        if category_dir.exists():
            for existing_file in category_dir.iterdir():
                if existing_file.is_file() and existing_file.suffix.lower() in self.DOC_EXTENSIONS:
                    # 排除当前正在处理的文件（避免误判）
                    if existing_file.name not in processing_filenames:
                        existing_files.add(existing_file.name)
        
        # 4. 遍历排序后的列表，只检查相邻文件是否有重复（优化性能）
        resolved_names = {}
        used_names = set()  # 已使用的文件名集合
        
        i = 0
        while i < len(file_info_list):
            new_filename, file_path = file_info_list[i]
            
            # 检查当前文件名是否与已存在的文件冲突
            if new_filename in existing_files:
                # 与已存在文件冲突，需要添加序号
                base_name = Path(new_filename).stem
                suffix = Path(new_filename).suffix
                counter = 1
                candidate_name = f"{base_name}_{counter}{suffix}"
                while candidate_name in existing_files or candidate_name in used_names:
                    counter += 1
                    candidate_name = f"{base_name}_{counter}{suffix}"
                resolved_names[file_path] = candidate_name
                used_names.add(candidate_name)
                logger.warning(f"文件名与已存在文件冲突，添加序号: {new_filename} -> {candidate_name}")
                i += 1
                continue
            
            # 检查后续相邻文件是否有重复（排序后重复文件会聚集在一起）
            # 由于已排序，只需要检查相邻的文件即可，遇到不同的文件名就可以停止
            duplicate_count = 1
            duplicate_files = [(file_path, new_filename)]
            
            # 检查后续相邻文件，直到遇到不同的文件名（排序后重复文件会连续出现）
            for j in range(i + 1, len(file_info_list)):
                next_filename, next_file_path = file_info_list[j]
                if next_filename == new_filename:
                    duplicate_count += 1
                    duplicate_files.append((next_file_path, next_filename))
                else:
                    # 由于已排序，如果不同则后续不会有重复，可以停止检查
                    break
            
            if duplicate_count == 1:
                # 没有重复，直接使用
                if new_filename not in used_names:
                    resolved_names[file_path] = new_filename
                    used_names.add(new_filename)
                else:
                    # 与已分配的名称冲突，添加序号
                    base_name = Path(new_filename).stem
                    suffix = Path(new_filename).suffix
                    counter = 1
                    candidate_name = f"{base_name}_{counter}{suffix}"
                    while candidate_name in existing_files or candidate_name in used_names:
                        counter += 1
                        candidate_name = f"{base_name}_{counter}{suffix}"
                    resolved_names[file_path] = candidate_name
                    used_names.add(candidate_name)
                    logger.info(f"文件名与已分配名称冲突，添加序号: {new_filename} -> {candidate_name}")
            else:
                # 有重复，需要添加序号区分
                base_name = Path(new_filename).stem
                suffix = Path(new_filename).suffix
                counter = 0
                
                for dup_file_path, dup_filename in duplicate_files:
                    if counter == 0:
                        # 第一个文件保持原名（如果可用）
                        if new_filename not in used_names:
                            resolved_names[dup_file_path] = new_filename
                            used_names.add(new_filename)
                        else:
                            counter = 1
                            candidate_name = f"{base_name}_{counter}{suffix}"
                            while candidate_name in existing_files or candidate_name in used_names:
                                counter += 1
                                candidate_name = f"{base_name}_{counter}{suffix}"
                            resolved_names[dup_file_path] = candidate_name
                            used_names.add(candidate_name)
                            logger.info(f"文件名重复，添加序号: {new_filename} -> {candidate_name}")
                    else:
                        # 后续文件添加序号
                        candidate_name = f"{base_name}_{counter}{suffix}"
                        while candidate_name in existing_files or candidate_name in used_names:
                            counter += 1
                            candidate_name = f"{base_name}_{counter}{suffix}"
                        resolved_names[dup_file_path] = candidate_name
                        used_names.add(candidate_name)
                        logger.info(f"文件名重复，添加序号: {new_filename} -> {candidate_name}")
                    counter += 1
            
            i += duplicate_count
        
        return resolved_names
    
    def rename_file_worker(self, file_path: Path, new_filename: str) -> Tuple[bool, str, str]:
        """
        文件重命名工作函数（用于并行处理）
        
        Args:
            file_path: 原文件路径
            new_filename: 新文件名
            
        Returns:
            (是否成功, 旧文件名不含扩展名, 新文件名不含扩展名)
        """
        try:
            old_filename = file_path.name
            old_filename_without_ext = Path(old_filename).stem
            new_filename_without_ext = Path(new_filename).stem
            
            # 快速检查：只检查目标文件是否已存在（不检查整个目录）
            new_path = file_path.parent / new_filename
            if new_path.exists() and new_path != file_path:
                logger.warning(f"目标文件已存在，跳过重命名: {new_path}")
                return (False, old_filename_without_ext, new_filename_without_ext)
            
            if self.rename_file(file_path, new_filename):
                return (True, old_filename_without_ext, new_filename_without_ext)
            else:
                return (False, old_filename_without_ext, new_filename_without_ext)
        except Exception as e:
            logger.error(f"重命名文件失败: {file_path}, 错误: {e}")
            return (False, "", "")
    
    def process_files(self, files_by_category: Dict[str, List[Tuple[Path, str]]]):
        """
        处理文件：重命名并批量更新Excel（优化版本）
        
        Args:
            files_by_category: 按分类分组的文件字典
        """
        # 按分类处理，每个分类的Excel文件只打开一次
        for category, files_list in files_by_category.items():
            if not files_list:
                continue
            
            logger.info(f"处理分类 '{category}'，共 {len(files_list)} 个文件")
            
            # 1. 检查并解决重复文件名（排序后只检查相邻文件，优化性能）
            category_dir = self.output_dir / category
            logger.info(f"检查文件名重复情况...")
            resolved_names = self.check_and_resolve_duplicates(files_list, category_dir)
            
            # 2. 并行重命名文件（文件系统操作可以并行）
            rename_results = []
            with ThreadPoolExecutor(max_workers=10) as executor:
                future_to_file = {}
                for file_path, _ in files_list:
                    if file_path not in resolved_names:
                        logger.warning(f"文件未在解析列表中，跳过: {file_path}")
                        continue
                    
                    new_filename = resolved_names[file_path]
                    old_filename = file_path.name
                    future = executor.submit(self.rename_file_worker, file_path, new_filename)
                    future_to_file[future] = (file_path, old_filename, new_filename)
                
                for future in as_completed(future_to_file):
                    file_path, old_filename, new_filename = future_to_file[future]
                    try:
                        success, old_name_no_ext, new_name_no_ext = future.result()
                        if success:
                            self.stats['files_renamed'] += 1
                            rename_results.append((old_name_no_ext, new_name_no_ext))
                            logger.info(f"文件重命名成功: {old_filename} -> {new_filename}")
                        else:
                            self.stats['errors'] += 1
                    except Exception as e:
                        logger.error(f"处理文件失败: {file_path}, 错误: {e}")
                        self.stats['errors'] += 1
            
            # 3. 批量更新Excel（每个分类的Excel文件只打开一次）
            if rename_results:
                excel_file = self.output_dir / self.CATEGORY_TO_EXCEL[category]
                filename_mappings = {old: new for old, new in rename_results}
                updated_rows = self.batch_update_excel_filenames(excel_file, filename_mappings)
                self.stats['excel_rows_updated'] += updated_rows
    
    def print_summary(self):
        """打印统计信息"""
        print("\n" + "=" * 60)
        print("文件名规范化完成统计")
        print("=" * 60)
        print(f"扫描文件总数: {self.stats['total_files']}")
        print(f"重命名文件数量: {self.stats['files_renamed']}")
        print(f"更新Excel行数: {self.stats['excel_rows_updated']}")
        print(f"错误数量: {self.stats['errors']}")
        print("=" * 60)
    
    def run(self):
        """运行规范化流程"""
        try:
            print("=" * 60)
            print("文件名规范化脚本")
            print("=" * 60)
            print(f"输出目录: {self.output_dir}")
            print()
            print("规范化规则:")
            print("  1. 空格、Tab等空白字符 -> 下划线 (_)")
            print("  2. 英文圆括号 () -> 中文圆括号 （）")
            print()
            
            # 1. 扫描文件
            print("步骤1: 扫描文件...")
            files_by_category = self.scan_files()
            total_files_to_process = sum(len(files) for files in files_by_category.values())
            print(f"扫描完成，共找到 {self.stats['total_files']} 个文件")
            print(f"需要规范化的文件: {total_files_to_process} 个")
            print()
            
            if total_files_to_process == 0:
                print("没有需要规范化的文件，退出")
                return
            
            # 2. 显示将要处理的文件列表（只显示前50个，避免输出过多）
            print("将要处理的文件列表（显示前50个）:")
            display_count = 0
            for category, files_list in files_by_category.items():
                for file_path, _ in files_list:
                    if display_count >= 50:
                        break
                    old_filename = file_path.name
                    new_filename = self.normalize_filename(old_filename)
                    print(f"  [{category}] {old_filename} -> {new_filename}")
                    display_count += 1
                if display_count >= 50:
                    break
            if total_files_to_process > 50:
                print(f"  ... 还有 {total_files_to_process - 50} 个文件未显示")
            print()
            
            # 3. 确认操作
            response = input(f"是否继续处理 {total_files_to_process} 个文件？(y/n): ").strip().lower()
            
            if response != 'y':
                print("操作已取消")
                return
            
            print()
            
            # 4. 处理文件（批量处理，优化性能）
            print("步骤2: 重命名文件并批量更新Excel...")
            self.process_files(files_by_category)
            print()
            
            # 5. 打印统计信息
            self.print_summary()
            
        except KeyboardInterrupt:
            print("\n\n程序被用户中断")
            sys.exit(0)
        except Exception as e:
            logger.error(f"程序运行失败: {e}", exc_info=True)
            print(f"\n错误: {e}")
            sys.exit(1)


def main():
    """主函数"""
    normalizer = FilenameNormalizer()
    normalizer.run()


if __name__ == "__main__":
    main()

